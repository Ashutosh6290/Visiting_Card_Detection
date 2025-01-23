from flask import Flask, render_template, request, send_file
import os
from PIL import Image as PILImage
import pytesseract
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import re

app = Flask(__name__)

# Set the path to the Tesseract executable
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'  # Adjust this path

def extract_text_from_image(image_path):
    """Extract text from a single image using Tesseract OCR."""
    try:
        image = PILImage.open(image_path)
        text = pytesseract.image_to_string(image)
        return text
    except Exception as e:
        print(f"Error processing image {image_path}: {e}")
        return ""

def parse_text(text):
    """Parse extracted text into specific fields."""
    data = {
        "Name": None,
        "Designation": None,
        "Phone Number": None,
        "Email": None,
        "Website": None,
        "Address": None,
        "Company Name": None,
        "Other Information": None
    }

    phone_pattern = r'\b\d{10}\b|\(\d{3}\)\s?\d{3}-\d{4}|\+\d{1,3}\s?\d{10}'
    email_pattern = r'\b[A-Za-z0-9._%+-]+@([A-Za-z0-9.-]+)\.(com|in|org|net)\b'
    website_pattern = r'\b(?:http://|https://)?(?:www\.)?[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}\b'
    designation_keywords = r'\b(Manager|Founder|Engineer|Developer|Executive|Director|Officer|Specialist|Consultant|Coordinator|Analyst|Technician|Designer|Assistant|Head|Lead|Senior|Sales)\b'

    lines = text.splitlines()
    for i, line in enumerate(lines):
        line = line.strip()
        if re.search(designation_keywords, line, re.IGNORECASE):
            data["Designation"] = line
            if i > 0:
                data["Name"] = lines[i - 1].strip()
            break

    phone_match = re.search(phone_pattern, text)
    if phone_match:
        data["Phone Number"] = phone_match.group(0)

    email_match = re.search(email_pattern, text)
    if email_match:
        data["Email"] = email_match.group(0)
        data["Company Name"] = email_match.group(1)

    website_match = re.search(website_pattern, text)
    if website_match:
        data["Website"] = website_match.group(0)

    data["Other Information"] = text.strip()

    return data

def process_folder(folder_path, output_excel):
    common_image_folder = os.path.join(folder_path, "Processed_Images")
    if not os.path.exists(common_image_folder):
        os.makedirs(common_image_folder)

    wb = Workbook()
    ws = wb.active
    ws.title = "Visiting Card Details"
    ws.append([
        "Filename", "Image Link", "Name", "Designation", "Phone Number",
        "Email", "Company Name", "Website", "Other Information"
    ])

    for row_index, filename in enumerate(os.listdir(folder_path), start=2):
        if filename.lower().endswith((".png", ".jpg", ".jpeg", ".tiff", ".bmp")):
            source_file_path = os.path.join(folder_path, filename)
            dest_file_path = os.path.join(common_image_folder, filename)

            if not os.path.exists(dest_file_path):
                with open(source_file_path, "rb") as src, open(dest_file_path, "wb") as dst:
                    dst.write(src.read())

            extracted_text = extract_text_from_image(source_file_path)
            parsed_data = parse_text(extracted_text)

            ws.append([
                filename, dest_file_path, parsed_data["Name"], parsed_data["Designation"],
                parsed_data["Phone Number"], parsed_data["Email"], parsed_data["Company Name"],
                parsed_data["Website"], parsed_data["Other Information"]
            ])

            ws[f"B{row_index}"].hyperlink = dest_file_path
            ws[f"B{row_index}"].value = "Open Image"
            ws[f"B{row_index}"].font = Font(color="0000FF", underline="single")

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 5

    wb.save(output_excel)

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        folder_path = request.form["folder_path"]
        output_excel = request.form["output_file"]

        if os.path.isdir(folder_path):
            process_folder(folder_path, output_excel)
            return render_template("index.html", success=True, download_link=output_excel)
        else:
            return render_template("index.html", error="Invalid folder path.")

    return render_template("index.html")

@app.route("/download/<filename>")
def download_file(filename):
    return send_file(filename, as_attachment=True)

if __name__ == "__main__":
    app.run(debug=True)
